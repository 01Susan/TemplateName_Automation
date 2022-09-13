from openpyxl import Workbook, load_workbook
import xlsxwriter as writer
from  alive_progress import alive_bar

# file_path = r"C:\Users\susan_ksr4b\Documents\pythonProject\
# GEBRUDER WISE 2.0 Work Sheet.xlsx"

# Asking the user for the file path
file_path = input("Enter the file path of the document: ").replace('"', '')
# Asking the user for the column name
column_name = input("Enter the column name that you want to extract: ").upper()
list_name = set()
load_workbook = load_workbook(file_path)
work_sheet = load_workbook.active

for row in work_sheet[column_name]:
    values = row.value
    if values != None:
        list_name.add(values.lower())

# Creating the new sheet

new_workbook = writer.Workbook("Template Name.xlsx")
new_worksheet = new_workbook.add_worksheet()

row, col = 0, 0
title = "Running"
with alive_bar(len(list_name), title=title, dual_line=True) as bar:
    for element in list_name:
        bar.text = f"Writing {element},Please Wait"
        new_worksheet.write(row, col, element)
        row += 1
        bar()
print("Done👍")
new_workbook.close()
